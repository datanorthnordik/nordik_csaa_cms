import sys
from collections import defaultdict
from openpyxl import load_workbook

def norm_rgb(rgb: str) -> str:
    """
    openpyxl often returns ARGB like 'FFRRGGBB'.
    Normalize to '#RRGGBB' (drop alpha if present).
    """
    if not rgb:
        return ""
    rgb = rgb.strip()
    # Sometimes it's already 'RRGGBB' or 'FFRRGGBB'
    if len(rgb) == 8:  # ARGB
        rgb = rgb[2:]
    if len(rgb) == 6:
        return "#" + rgb.upper()
    return rgb.upper()

def extract_unique_fills(xlsx_path: str, sheet_name: str | None = None, include_samples: bool = True):
    wb = load_workbook(xlsx_path, data_only=True)

    sheets = [wb[sheet_name]] if sheet_name else wb.worksheets

    unique_colors = set()
    samples = defaultdict(list)  # color -> list of "Sheet!A1"

    for ws in sheets:
        for row in ws.iter_rows():
            for cell in row:
                fill = cell.fill
                if not fill:
                    continue

                # We only care about solid/pattern fills that carry a visible color
                # fill.patternType can be None, 'solid', 'gray125', etc.
                # For "colored cells", most of the time patternType == 'solid'
                fg = None
                if fill.patternType:  # pattern fill
                    fg = fill.fgColor

                if fg is None:
                    continue

                # openpyxl color types: 'rgb', 'indexed', 'theme'
                color_key = None

                if fg.type == "rgb" and fg.rgb:
                    color_key = norm_rgb(fg.rgb)

                elif fg.type == "indexed" and fg.indexed is not None:
                    # Indexed colors exist, but mapping depends on workbook palette.
                    # We keep as a stable identifier.
                    color_key = f"INDEXED:{fg.indexed}"

                elif fg.type == "theme" and fg.theme is not None:
                    # Theme colors depend on theme + tint; keep identifier
                    tint = fg.tint if fg.tint is not None else 0
                    color_key = f"THEME:{fg.theme}:TINT:{tint}"

                if color_key:
                    unique_colors.add(color_key)
                    if include_samples and len(samples[color_key]) < 5:
                        samples[color_key].append(f"{ws.title}!{cell.coordinate}")

    # Print results
    print("\nUnique fill colors found:")
    for c in sorted(unique_colors):
        print(" -", c)

    if include_samples:
        print("\nSample locations (up to 5 per color):")
        for c in sorted(unique_colors):
            locs = ", ".join(samples[c])
            print(f" - {c}: {locs}")

    print(f"\nTotal unique fills: {len(unique_colors)}")

def main():
    if len(sys.argv) < 2:
        print("Usage: python extract_excel_colors.py <path-to-xlsx> [sheetName]")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    sheet_name = sys.argv[2] if len(sys.argv) >= 3 else None

    extract_unique_fills(xlsx_path, sheet_name=sheet_name, include_samples=True)

if __name__ == "__main__":
    main()
